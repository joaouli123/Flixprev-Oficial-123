from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Callable

import pdfplumber
from docx import Document as DocxDocument
from openpyxl import load_workbook

try:
    import fitz
except Exception:  # pragma: no cover
    fitz = None


@dataclass(slots=True)
class ParsedSection:
    text: str
    source_file: str | None
    source_uri: str | None
    page_number: int | None
    section_title: str | None


class DocumentParser:
    """Parse local documents into normalized text sections."""

    @staticmethod
    def _normalize_text_block(text: str) -> str:
        lines: list[str] = []
        for raw_line in str(text or "").replace("\r\n", "\n").replace("\r", "\n").split("\n"):
            line = " ".join(str(raw_line or "").split())
            if line:
                lines.append(line)
            elif lines and lines[-1] != "":
                lines.append("")
        return "\n".join(lines).strip()

    @staticmethod
    def _normalize_table_cell(value: object) -> str:
        return " ".join(str(value or "").replace("\r\n", "\n").replace("\r", "\n").split())

    @staticmethod
    def _escape_markdown_cell(value: str) -> str:
        return str(value or "").replace("|", "\\|")

    @staticmethod
    def _row_looks_like_header(row: list[str]) -> bool:
        non_empty = [cell for cell in row if cell]
        if not non_empty:
            return False

        alpha_cells = sum(any(char.isalpha() for char in cell) for cell in non_empty)
        long_cells = sum(len(cell) > 80 for cell in non_empty)
        return alpha_cells >= max(1, len(non_empty) // 2) and long_cells == 0

    def _extract_pdf_page_text(self, fitz_page, plumber_page) -> str:
        fitz_text = self._normalize_text_block((fitz_page.get_text("text") or "") if fitz_page is not None else "")
        if fitz_text:
            return fitz_text
        return self._normalize_text_block((plumber_page.extract_text() or "") if plumber_page is not None else "")

    def _format_pdf_table(self, table_rows: list[list[object]], *, page_number: int, table_index: int) -> str:
        normalized_rows: list[list[str]] = []
        max_columns = 0

        for raw_row in table_rows or []:
            row = [self._normalize_table_cell(cell) for cell in (raw_row or [])]
            if not any(row):
                continue
            normalized_rows.append(row)
            max_columns = max(max_columns, len(row))

        if not normalized_rows or max_columns == 0:
            return ""

        padded_rows = [row + [""] * (max_columns - len(row)) for row in normalized_rows]
        generic_headers = [f"Column {index + 1}" for index in range(max_columns)]

        header_like = len(padded_rows) >= 2 and self._row_looks_like_header(padded_rows[0])
        markdown_headers = [cell or generic_headers[index] for index, cell in enumerate(padded_rows[0])] if header_like else generic_headers
        markdown_rows = padded_rows[1:] if header_like else padded_rows

        markdown_lines = [f"[TABLE {table_index} | PAGE {page_number}]"]
        if header_like:
            markdown_lines.append(f"Possible header row: {' | '.join(self._escape_markdown_cell(cell or generic_headers[index]) for index, cell in enumerate(padded_rows[0]))}")
        markdown_lines.extend(
            [
                f"| {' | '.join(self._escape_markdown_cell(cell) for cell in markdown_headers)} |",
                f"| {' | '.join('---' for _ in markdown_headers)} |",
            ]
        )
        for row in markdown_rows:
            markdown_lines.append(f"| {' | '.join(self._escape_markdown_cell(cell) for cell in row)} |")

        row_headers = markdown_headers if header_like else generic_headers
        structured_rows: list[str] = []
        for row_number, row in enumerate(markdown_rows, start=1 if header_like else 0):
            assignments = [
                f"{header} = {value}"
                for header, value in zip(row_headers, row)
                if value
            ]
            if assignments:
                structured_rows.append(f"- Row {row_number + 1}: {'; '.join(assignments)}")

        parts = ["\n".join(markdown_lines)]
        if structured_rows:
            parts.append("Structured rows:\n" + "\n".join(structured_rows))
        return "\n\n".join(part for part in parts if part.strip())

    def _extract_pdf_tables(self, plumber_page, *, page_number: int) -> list[str]:
        if plumber_page is None:
            return []

        table_candidates = plumber_page.extract_tables() or []
        if not table_candidates:
            table_candidates = plumber_page.extract_tables(
                table_settings={
                    "vertical_strategy": "text",
                    "horizontal_strategy": "text",
                    "intersection_tolerance": 5,
                    "snap_tolerance": 3,
                }
            ) or []

        formatted_tables: list[str] = []
        seen_tables: set[str] = set()
        for table_index, table_rows in enumerate(table_candidates, start=1):
            formatted = self._format_pdf_table(table_rows, page_number=page_number, table_index=table_index)
            if not formatted:
                continue
            signature = " ".join(formatted.split())
            if signature in seen_tables:
                continue
            seen_tables.add(signature)
            formatted_tables.append(formatted)
        return formatted_tables

    def parse_file(
        self,
        file_path: Path,
        *,
        progress_cb: Callable[[int, int, str], None] | None = None,
    ) -> list[ParsedSection]:
        extension = file_path.suffix.lower()
        if extension == ".pdf":
            return self._parse_pdf(file_path, progress_cb=progress_cb)
        if extension == ".docx":
            return self._parse_docx(file_path)
        if extension in {".xls", ".xlsx"}:
            return self._parse_excel(file_path)
        if extension in {".txt", ".md", ".csv"}:
            return self._parse_text(file_path)
        raise ValueError(f"Unsupported file extension: {extension}")

    def _parse_pdf(
        self,
        file_path: Path,
        *,
        progress_cb: Callable[[int, int, str], None] | None = None,
    ) -> list[ParsedSection]:
        sections: list[ParsedSection] = []
        fitz_doc = fitz.open(file_path) if fitz is not None else None

        try:
            with pdfplumber.open(file_path) as pdf:
                plumber_total_pages = len(pdf.pages)
                fitz_total_pages = int(getattr(fitz_doc, "page_count", 0) or 0)
                total_pages = max(plumber_total_pages, fitz_total_pages)

                for page_index in range(total_pages):
                    page_number = page_index + 1
                    plumber_page = pdf.pages[page_index] if page_index < plumber_total_pages else None
                    fitz_page = fitz_doc.load_page(page_index) if fitz_doc is not None and page_index < fitz_total_pages else None

                    text = self._extract_pdf_page_text(fitz_page, plumber_page)
                    if text:
                        sections.append(
                            ParsedSection(
                                text=text,
                                source_file=file_path.name,
                                source_uri=None,
                                page_number=page_number,
                                section_title=f"Page {page_number}",
                            )
                        )

                    for table_index, table_text in enumerate(self._extract_pdf_tables(plumber_page, page_number=page_number), start=1):
                        sections.append(
                            ParsedSection(
                                text=table_text,
                                source_file=file_path.name,
                                source_uri=None,
                                page_number=page_number,
                                section_title=f"Page {page_number} - Table {table_index}",
                            )
                        )

                    if progress_cb and total_pages > 0:
                        if page_number == 1 or page_number % 25 == 0 or page_number == total_pages:
                            progress_cb(page_number, total_pages, "Parsing PDF")
        finally:
            if fitz_doc is not None:
                fitz_doc.close()

        return sections

    def _parse_docx(self, file_path: Path) -> list[ParsedSection]:
        doc = DocxDocument(file_path)
        sections: list[ParsedSection] = []

        current_title = "Document"
        buffer: list[str] = []

        def flush() -> None:
            if not buffer:
                return
            text = "\n".join(buffer).strip()
            if text:
                sections.append(
                    ParsedSection(
                        text=text,
                        source_file=file_path.name,
                        source_uri=None,
                        page_number=None,
                        section_title=current_title,
                    )
                )
            buffer.clear()

        for paragraph in doc.paragraphs:
            text = paragraph.text.strip()
            if not text:
                continue
            style_name = (paragraph.style.name or "").lower() if paragraph.style else ""
            if "heading" in style_name:
                flush()
                current_title = text
                continue
            buffer.append(text)

        flush()
        return sections

    def _parse_excel(self, file_path: Path) -> list[ParsedSection]:
        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        sections: list[ParsedSection] = []

        try:
            for sheet in workbook.worksheets:
                lines: list[str] = []
                for row in sheet.iter_rows(values_only=True):
                    values = ["" if value is None else str(value).strip() for value in row]
                    if not any(values):
                        continue
                    lines.append(" | ".join(values))

                text = "\n".join(lines).strip()
                if not text:
                    continue

                sections.append(
                    ParsedSection(
                        text=text,
                        source_file=file_path.name,
                        source_uri=None,
                        page_number=None,
                        section_title=f"Sheet: {sheet.title}",
                    )
                )
        finally:
            workbook.close()

        return sections

    def _parse_text(self, file_path: Path) -> list[ParsedSection]:
        raw = file_path.read_bytes()
        decoded = None
        for encoding in ("utf-8", "utf-8-sig", "latin-1", "cp1252"):
            try:
                decoded = raw.decode(encoding)
                break
            except UnicodeDecodeError:
                continue

        if decoded is None:
            decoded = raw.decode("utf-8", errors="ignore")

        text = decoded.strip()
        if not text:
            return []

        return [
            ParsedSection(
                text=text,
                source_file=file_path.name,
                source_uri=None,
                page_number=None,
                section_title="Text File",
            )
        ]
